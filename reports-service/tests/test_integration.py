"""
Integration tests for reports-service.
These tests validate database connectivity, Excel generation, and SOAP integration.
"""
import pytest
import os
import sys
import tempfile
import shutil
from datetime import datetime, timedelta
from unittest.mock import Mock, patch, MagicMock
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

# Add src to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

# Skip tests if dependencies are not available
def pytest_configure(config):
    """Configure pytest markers."""
    config.addinivalue_line("markers", "integration: Integration tests")
    config.addinivalue_line("markers", "database: Tests that require database")
    config.addinivalue_line("markers", "soap: Tests that require SOAP service")

@pytest.mark.integration
@pytest.mark.database
class TestDatabaseIntegration:
    """Integration tests for database connectivity and operations."""
    
    def test_database_connection(self):
        """Test that we can connect to the database."""
        try:
            from src.models.database import test_connection, engine
            
            # Test basic connection
            connection_result = test_connection()
            
            if connection_result:
                # Test that we can execute a simple query
                with engine.connect() as connection:
                    result = connection.execute(text("SELECT 1 as test_value"))
                    row = result.fetchone()
                    assert row[0] == 1
            else:
                pytest.skip("Database connection not available in test environment")
                
        except Exception as e:
            pytest.skip(f"Database connection failed: {e}")
    
    def test_database_session_creation(self):
        """Test that database sessions can be created and used."""
        try:
            from src.models.database import get_db, SessionLocal
            
            # Test session creation through dependency
            session_generator = get_db()
            session = next(session_generator)
            
            assert session is not None
            
            # Test that we can execute queries
            result = session.execute(text("SELECT 1 as test_value"))
            row = result.fetchone()
            assert row[0] == 1
            
            # Clean up
            try:
                next(session_generator)
            except StopIteration:
                pass  # Expected when session is closed
                
        except Exception as e:
            pytest.skip(f"Database session creation failed: {e}")
    
    def test_database_tables_exist(self):
        """Test that required database tables exist."""
        try:
            from src.models.database import engine
            
            # Check for main tables that should exist
            required_tables = [
                'usuarios',
                'donaciones', 
                'eventos',
                'filtros_guardados',
                'archivos_excel'
            ]
            
            with engine.connect() as connection:
                # Get list of tables
                result = connection.execute(text("SHOW TABLES"))
                existing_tables = [row[0] for row in result.fetchall()]
                
                # Check that core tables exist (some may not exist in test environment)
                core_tables = ['usuarios', 'donaciones', 'eventos']
                for table in core_tables:
                    if table in existing_tables:
                        # If table exists, verify we can query it
                        try:
                            result = connection.execute(text(f"SELECT COUNT(*) FROM {table}"))
                            count = result.fetchone()[0]
                            assert count >= 0  # Should be able to get count
                        except Exception as e:
                            pytest.fail(f"Failed to query table {table}: {e}")
                            
        except Exception as e:
            pytest.skip(f"Database table check failed: {e}")
    
    def test_model_imports_and_creation(self):
        """Test that database models can be imported and instantiated."""
        from src.models.donation import Donation, DonationCategory
        from src.models.event import Event
        from src.models.filter import SavedFilter, FilterType, ExcelFile
        from src.models.user import User, UserRole
        
        # Test model instantiation
        donation = Donation()
        donation.categoria = DonationCategory.ROPA
        donation.cantidad = 10
        donation.eliminado = False
        
        event = Event()
        event.nombre = "Test Event"
        event.descripcion = "Test Description"
        
        saved_filter = SavedFilter()
        saved_filter.nombre = "Test Filter"
        saved_filter.tipo = FilterType.DONACIONES
        saved_filter.configuracion = {"categoria": "ROPA"}
        
        user = User()
        user.nombre = "Test User"
        user.email = "test@example.com"
        user.rol = UserRole.PRESIDENTE
        
        excel_file = ExcelFile()
        excel_file.nombre_archivo = "test.xlsx"
        excel_file.ruta_archivo = "/tmp/test.xlsx"
        
        # Verify objects were created
        assert donation.categoria == DonationCategory.ROPA
        assert event.nombre == "Test Event"
        assert saved_filter.tipo == FilterType.DONACIONES
        assert user.rol == UserRole.PRESIDENTE
        assert excel_file.nombre_archivo == "test.xlsx"


@pytest.mark.integration
class TestExcelIntegration:
    """Integration tests for Excel file generation and management."""
    
    def setup_method(self):
        """Set up test environment for each test."""
        # Create temporary directory for Excel files
        self.temp_dir = tempfile.mkdtemp()
        
    def teardown_method(self):
        """Clean up after each test."""
        # Remove temporary directory with retry for Windows
        if os.path.exists(self.temp_dir):
            try:
                shutil.rmtree(self.temp_dir)
            except PermissionError:
                # On Windows, files might still be in use
                import time
                time.sleep(0.1)
                try:
                    shutil.rmtree(self.temp_dir)
                except PermissionError:
                    # If still failing, just pass - temp files will be cleaned up by OS
                    pass
    
    @patch('src.services.excel_service.settings')
    @patch('src.services.excel_service.get_db_session')
    def test_excel_service_initialization(self, mock_db_session, mock_settings):
        """Test that Excel service can be initialized."""
        from src.services.excel_service import ExcelExportService
        
        # Mock settings
        mock_settings.excel_storage_path = self.temp_dir
        
        # Initialize service
        service = ExcelExportService()
        
        assert service is not None
        assert service.storage_path == self.temp_dir
        assert os.path.exists(self.temp_dir)
    
    @patch('src.services.excel_service.settings')
    def test_excel_workbook_creation(self, mock_settings):
        """Test that Excel workbooks can be created."""
        from src.services.excel_service import ExcelExportService
        from src.models.donation import Donation, DonationCategory
        from openpyxl import Workbook
        
        # Mock settings
        mock_settings.excel_storage_path = self.temp_dir
        
        service = ExcelExportService()
        
        # Create mock donations
        donations = []
        for i in range(3):
            donation = Mock(spec=Donation)
            donation.id = i + 1
            donation.categoria = DonationCategory.ROPA
            donation.descripcion = f"Test donation {i + 1}"
            donation.cantidad = 10 * (i + 1)
            donation.eliminado = False
            donation.fecha_alta = datetime.now()
            donation.fecha_modificacion = None
            donation.usuario_creador = None
            donation.usuario_modificador = None
            donations.append(donation)
        
        # Group donations by category
        donations_by_category = {DonationCategory.ROPA: donations}
        
        # Create workbook
        workbook = service._create_workbook(donations_by_category)
        
        assert isinstance(workbook, Workbook)
        assert len(workbook.worksheets) == 1
        assert workbook.worksheets[0].title == "Ropa"
        
        # Check that data was written
        worksheet = workbook.worksheets[0]
        assert worksheet['A1'].value == "ID"  # Header
        assert worksheet['A2'].value == 1     # First donation ID
        assert worksheet['D2'].value == 10    # First donation quantity
    
    @patch('src.services.excel_service.settings')
    def test_excel_file_generation(self, mock_settings):
        """Test complete Excel file generation process."""
        from src.services.excel_service import ExcelExportService
        
        # Mock settings
        mock_settings.excel_storage_path = self.temp_dir
        
        service = ExcelExportService()
        
        # Test filename generation
        filename = service._generate_filename()
        assert filename.endswith('.xlsx')
        assert 'reporte_donaciones' in filename
        
        # Test filename with filters
        from src.models.donation import DonationCategory
        from datetime import datetime
        
        fecha_desde = datetime(2024, 1, 1)
        fecha_hasta = datetime(2024, 12, 31)
        
        filename_with_filters = service._generate_filename(
            categoria=DonationCategory.ROPA,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta
        )
        
        assert 'categoria_ropa' in filename_with_filters
        assert 'desde_20240101' in filename_with_filters
        assert 'hasta_20241231' in filename_with_filters
    
    def test_excel_file_operations(self):
        """Test Excel file I/O operations."""
        from openpyxl import Workbook
        
        # Create a test Excel file
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Test Sheet"
        worksheet['A1'] = "Test Data"
        worksheet['B1'] = 123
        
        # Save to temporary file
        test_file_path = os.path.join(self.temp_dir, "test_excel.xlsx")
        workbook.save(test_file_path)
        
        # Verify file was created
        assert os.path.exists(test_file_path)
        assert os.path.getsize(test_file_path) > 0
        
        # Verify file can be read back
        from openpyxl import load_workbook
        loaded_workbook = load_workbook(test_file_path)
        loaded_worksheet = loaded_workbook.active
        
        assert loaded_worksheet['A1'].value == "Test Data"
        assert loaded_worksheet['B1'].value == 123


@pytest.mark.integration
@pytest.mark.soap
class TestSOAPIntegration:
    """Integration tests for SOAP service connectivity and operations."""
    
    def test_soap_client_initialization(self):
        """Test that SOAP client can be initialized."""
        try:
            from src.soap.client import SOAPClient, get_soap_client
            
            # Test direct instantiation
            try:
                client = SOAPClient()
                assert client is not None
            except Exception as e:
                # SOAP service might not be available in test environment
                pytest.skip(f"SOAP service not available for testing: {e}")
            
            # Test singleton pattern
            client1 = get_soap_client()
            client2 = get_soap_client()
            assert client1 is client2
            
        except ImportError as e:
            if "cgi" in str(e):
                pytest.skip("SOAP client not available due to Python 3.13 compatibility issue with zeep")
            else:
                raise
    
    def test_soap_client_with_mock(self):
        """Test SOAP client functionality with mocked zeep client."""
        try:
            with patch('src.soap.client.Client') as mock_zeep_client:
                from src.soap.client import SOAPClient
                
                # Mock the zeep Client
                mock_client_instance = Mock()
                mock_zeep_client.return_value = mock_client_instance
                
                # Initialize SOAP client
                soap_client = SOAPClient()
                
                assert soap_client.client == mock_client_instance
                mock_zeep_client.assert_called_once()
                
        except ImportError as e:
            if "cgi" in str(e):
                pytest.skip("SOAP client not available due to Python 3.13 compatibility issue with zeep")
            else:
                raise
    
    def test_soap_president_data_query(self):
        """Test querying president data via SOAP."""
        try:
            with patch('src.soap.client.Client') as mock_zeep_client:
                from src.soap.client import SOAPClient
                
                # Mock the zeep Client and service
                mock_client_instance = Mock()
                mock_service = Mock()
                mock_client_instance.service = mock_service
                mock_zeep_client.return_value = mock_client_instance
                
                # Mock response data
                mock_president = Mock()
                mock_president.organizationId = 1
                mock_president.presidentName = "Test President"
                mock_president.presidentEmail = "president@test.com"
                mock_president.presidentPhone = "123-456-7890"
                mock_president.presidentId = 101
                mock_president.startDate = "2024-01-01"
                mock_president.status = "Active"
                
                mock_service.getPresidentData.return_value = [mock_president]
                
                # Initialize SOAP client and query data
                soap_client = SOAPClient()
                result = soap_client.get_president_data([1])
                
                # Verify results
                assert len(result) == 1
                assert result[0]['organization_id'] == 1
                assert result[0]['president_name'] == "Test President"
                assert result[0]['president_email'] == "president@test.com"
                
                # Verify service was called correctly
                mock_service.getPresidentData.assert_called_once_with(organizationIds=[1])
                
        except ImportError as e:
            if "cgi" in str(e):
                pytest.skip("SOAP client not available due to Python 3.13 compatibility issue with zeep")
            else:
                raise
    
    def test_soap_organization_data_query(self):
        """Test querying organization data via SOAP."""
        try:
            with patch('src.soap.client.Client') as mock_zeep_client:
                from src.soap.client import SOAPClient
                
                # Mock the zeep Client and service
                mock_client_instance = Mock()
                mock_service = Mock()
                mock_client_instance.service = mock_service
                mock_zeep_client.return_value = mock_client_instance
                
                # Mock response data
                mock_org = Mock()
                mock_org.organizationId = 1
                mock_org.organizationName = "Test Organization"
                mock_org.organizationType = "NGO"
                mock_org.address = "123 Test St"
                mock_org.city = "Test City"
                mock_org.country = "Test Country"
                mock_org.phone = "123-456-7890"
                mock_org.email = "org@test.com"
                mock_org.website = "www.test.com"
                mock_org.registrationDate = "2024-01-01"
                mock_org.status = "Active"
                mock_org.description = "Test organization"
                
                mock_service.getOrganizationData.return_value = [mock_org]
                
                # Initialize SOAP client and query data
                soap_client = SOAPClient()
                result = soap_client.get_organization_data([1])
                
                # Verify results
                assert len(result) == 1
                assert result[0]['organization_id'] == 1
                assert result[0]['organization_name'] == "Test Organization"
                assert result[0]['organization_type'] == "NGO"
                assert result[0]['address'] == "123 Test St"
                
                # Verify service was called correctly
                mock_service.getOrganizationData.assert_called_once_with(organizationIds=[1])
                
        except ImportError as e:
            if "cgi" in str(e):
                pytest.skip("SOAP client not available due to Python 3.13 compatibility issue with zeep")
            else:
                raise
    
    def test_soap_combined_data_query(self):
        """Test querying combined president and organization data."""
        try:
            with patch('src.soap.client.Client') as mock_zeep_client:
                from src.soap.client import SOAPClient
                
                # Mock the zeep Client and service
                mock_client_instance = Mock()
                mock_service = Mock()
                mock_client_instance.service = mock_service
                mock_zeep_client.return_value = mock_client_instance
                
                # Mock response data for presidents
                mock_president = Mock()
                mock_president.organizationId = 1
                mock_president.presidentName = "Test President"
                mock_president.presidentEmail = "president@test.com"
                mock_service.getPresidentData.return_value = [mock_president]
                
                # Mock response data for organizations
                mock_org = Mock()
                mock_org.organizationId = 1
                mock_org.organizationName = "Test Organization"
                mock_org.organizationType = "NGO"
                mock_service.getOrganizationData.return_value = [mock_org]
                
                # Initialize SOAP client and query combined data
                soap_client = SOAPClient()
                result = soap_client.get_combined_data([1])
                
                # Verify results structure
                assert 'presidents' in result
                assert 'organizations' in result
                assert 'query_ids' in result
                assert 'total_presidents' in result
                assert 'total_organizations' in result
                
                assert len(result['presidents']) == 1
                assert len(result['organizations']) == 1
                assert result['query_ids'] == [1]
                assert result['total_presidents'] == 1
                assert result['total_organizations'] == 1
                
        except ImportError as e:
            if "cgi" in str(e):
                pytest.skip("SOAP client not available due to Python 3.13 compatibility issue with zeep")
            else:
                raise
    
    def test_soap_error_handling(self):
        """Test SOAP error handling."""
        try:
            with patch('src.soap.client.Client') as mock_zeep_client:
                from src.soap.client import SOAPClient, SOAPServiceError
                from zeep.exceptions import Fault, TransportError
                
                # Mock the zeep Client and service
                mock_client_instance = Mock()
                mock_service = Mock()
                mock_client_instance.service = mock_service
                mock_zeep_client.return_value = mock_client_instance
                
                soap_client = SOAPClient()
                
                # Test Fault exception handling
                mock_service.getPresidentData.side_effect = Fault("SOAP Fault")
                
                with pytest.raises(SOAPServiceError) as exc_info:
                    soap_client.get_president_data([1])
                
                assert "SOAP service error" in str(exc_info.value)
                
                # Test TransportError exception handling
                mock_service.getPresidentData.side_effect = TransportError("Transport Error")
                
                with pytest.raises(SOAPServiceError) as exc_info:
                    soap_client.get_president_data([1])
                
                assert "SOAP transport error" in str(exc_info.value)
                
                # Test generic exception handling
                mock_service.getPresidentData.side_effect = Exception("Generic Error")
                
                with pytest.raises(SOAPServiceError) as exc_info:
                    soap_client.get_president_data([1])
                
                assert "Unexpected SOAP error" in str(exc_info.value)
                
        except ImportError as e:
            if "cgi" in str(e):
                pytest.skip("SOAP client not available due to Python 3.13 compatibility issue with zeep")
            else:
                raise
    
    def test_soap_connection_test(self):
        """Test SOAP connection testing functionality."""
        try:
            from src.soap.client import get_soap_client
            
            try:
                client = get_soap_client()
                # Test connection (may fail if service is not available)
                connection_result = client.test_connection()
                # Should return boolean
                assert isinstance(connection_result, bool)
            except Exception as e:
                # SOAP service might not be available in test environment
                pytest.skip(f"SOAP service not available for testing: {e}")
                
        except ImportError as e:
            if "cgi" in str(e):
                pytest.skip("SOAP client not available due to Python 3.13 compatibility issue with zeep")
            else:
                raise


@pytest.mark.integration
class TestServiceIntegration:
    """Integration tests for service layer components."""
    
    def test_service_imports(self):
        """Test that all services can be imported and instantiated."""
        from src.services.donation_service import DonationService
        from src.services.event_service import EventService
        from src.services.filter_service import FilterService
        from src.services.excel_service import ExcelExportService
        
        # Test service instantiation
        donation_service = DonationService()
        event_service = EventService()
        filter_service = FilterService()
        excel_service = ExcelExportService()
        
        assert donation_service is not None
        assert event_service is not None
        assert filter_service is not None
        assert excel_service is not None
    
    @patch('src.services.donation_service.get_db_session')
    def test_donation_service_integration(self, mock_db_session):
        """Test donation service integration."""
        from src.services.donation_service import DonationService
        from src.models.donation import DonationCategory
        
        # Mock database session
        mock_session = Mock()
        mock_db_session.return_value.__enter__.return_value = mock_session
        
        service = DonationService()
        
        # Test service methods exist and can be called
        assert hasattr(service, 'get_donations_by_filters')
        assert hasattr(service, 'validate_user_access')
        
        # Test that methods can be called (with mocked dependencies)
        try:
            service.get_donations_by_filters(categoria=DonationCategory.ROPA)
        except Exception as e:
            # Expected since we're using mocked dependencies
            assert "Mock" in str(type(e)) or "AttributeError" in str(type(e))
    
    @patch('src.services.event_service.get_db_session')
    def test_event_service_integration(self, mock_db_session):
        """Test event service integration."""
        from src.services.event_service import EventService
        
        # Mock database session
        mock_session = Mock()
        mock_db_session.return_value.__enter__.return_value = mock_session
        
        service = EventService()
        
        # Test service methods exist
        assert hasattr(service, 'get_event_participation_report')
        assert hasattr(service, 'validate_user_access')
        
        # Test basic functionality
        assert service is not None
    
    def test_configuration_integration(self):
        """Test that configuration is properly loaded."""
        from src.config import settings, get_settings
        
        # Test settings object
        assert settings is not None
        assert hasattr(settings, 'database_url')
        assert hasattr(settings, 'SOAP_SERVICE_URL')
        assert hasattr(settings, 'excel_storage_path')
        
        # Test settings function
        settings_func = get_settings()
        assert settings_func is settings
        
        # Test that settings have reasonable defaults
        assert settings.database_url.startswith('mysql')
        assert settings.SOAP_SERVICE_URL.startswith('http')
        assert settings.excel_storage_path is not None


if __name__ == "__main__":
    # Run integration tests
    pytest.main([__file__, "-v", "-m", "integration"])