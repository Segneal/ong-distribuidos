"""
Excel export service for generating transfer reports in Excel format.
"""
import os
import uuid
from typing import List, Optional, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..models.transfer import DonationTransfer, TransferType, TransferStatus
from ..models.filter import ExcelFile
from ..models.user import User
from ..services.transfer_service import TransferService
from ..utils.database_utils import get_db_session
from ..config import settings


class TransferExcelService:
    """Service for generating Excel exports of transfer reports"""
    
    def __init__(self):
        self.transfer_service = TransferService()
        self.storage_path = settings.excel_storage_path
        
        # Ensure storage directory exists
        os.makedirs(self.storage_path, exist_ok=True)
    
    def generate_transfer_excel(
        self,
        user_id: int,
        tipo: Optional[TransferType] = None,
        fecha_desde: Optional[datetime] = None,
        fecha_hasta: Optional[datetime] = None,
        estado: Optional[TransferStatus] = None,
        user_organization: Optional[str] = None
    ) -> ExcelFile:
        """
        Generate Excel file with transfer data filtered by the provided criteria.
        
        Args:
            user_id: ID of the user requesting the export
            tipo: Filter by transfer type (optional)
            fecha_desde: Filter transfers from this date (optional)
            fecha_hasta: Filter transfers until this date (optional)
            estado: Filter by transfer status (optional)
            user_organization: User's organization (optional)
        
        Returns:
            ExcelFile object with file information
        """
        print("[TRANSFER EXCEL] Getting transfers from service...")
        # Get filtered transfers using the service
        transfers = self.transfer_service.get_transfers_by_filters(
            tipo=tipo,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            estado=estado,
            user_organization=user_organization
        )
        
        print(f"[TRANSFER EXCEL] Got {len(transfers)} transfers")
        
        # Group transfers by type and organization
        print("[TRANSFER EXCEL] Grouping transfers...")
        transfers_by_type_org = self._group_transfers_by_type_org(transfers)
        print(f"[TRANSFER EXCEL] Grouped into {len(transfers_by_type_org)} groups")
        
        # Generate Excel file
        file_id = str(uuid.uuid4())
        filename = self._generate_filename(tipo, fecha_desde, fecha_hasta)
        file_path = os.path.join(self.storage_path, f"{file_id}_{filename}")
        
        workbook = self._create_workbook(transfers_by_type_org)
        workbook.save(file_path)
        
        # Save file record to database
        with get_db_session() as session:
            excel_file = ExcelFile.create_with_expiration(
                usuario_id=user_id,
                nombre_archivo=filename,
                ruta_archivo=file_path,
                hours_to_expire=24  # Files expire after 24 hours
            )
            session.add(excel_file)
            session.commit()
            session.refresh(excel_file)
            
            return excel_file
    
    def _group_transfers_by_type_org(self, transfers: List[DonationTransfer]) -> Dict[str, List[DonationTransfer]]:
        """Group transfers by type and organization"""
        grouped = {}
        for transfer in transfers:
            # Create key with type and organization
            key = f"{transfer.tipo.value}_{transfer.organizacion_contraparte}"
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(transfer)
        return grouped
    
    def _generate_filename(
        self,
        tipo: Optional[TransferType] = None,
        fecha_desde: Optional[datetime] = None,
        fecha_hasta: Optional[datetime] = None
    ) -> str:
        """Generate descriptive filename for the Excel export"""
        parts = ["reporte_transferencias"]
        
        if tipo:
            parts.append(f"tipo_{tipo.value.lower()}")
        
        if fecha_desde:
            parts.append(f"desde_{fecha_desde.strftime('%Y%m%d')}")
        
        if fecha_hasta:
            parts.append(f"hasta_{fecha_hasta.strftime('%Y%m%d')}")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        parts.append(timestamp)
        
        return f"{'_'.join(parts)}.xlsx"
    
    def _create_workbook(self, transfers_by_type_org: Dict[str, List[DonationTransfer]]) -> Workbook:
        """Create Excel workbook with separate sheets for each type/organization"""
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create sheets for each group
        for group_key, transfers in transfers_by_type_org.items():
            sheet_name = self._get_sheet_name(group_key)
            worksheet = workbook.create_sheet(title=sheet_name)
            self._populate_worksheet(worksheet, transfers, group_key)
        
        # If no transfers, create an empty summary sheet
        if not transfers_by_type_org:
            worksheet = workbook.create_sheet(title="Sin Datos")
            worksheet['A1'] = "No se encontraron transferencias con los filtros aplicados"
            worksheet['A1'].font = Font(bold=True)
        
        return workbook
    
    def _get_sheet_name(self, group_key: str) -> str:
        """Get human-readable sheet name for group"""
        parts = group_key.split('_', 1)
        if len(parts) == 2:
            tipo, org = parts
            tipo_name = "Enviadas" if tipo == "ENVIADA" else "Recibidas"
            # Truncate organization name if too long
            org_name = org[:20] + "..." if len(org) > 20 else org
            return f"{tipo_name} - {org_name}"
        return group_key[:31]  # Excel sheet name limit
    
    def _populate_worksheet(self, worksheet, transfers: List[DonationTransfer], group_key: str):
        """Populate worksheet with transfer data"""
        # Define headers
        headers = [
            "ID",
            "Fecha",
            "Tipo",
            "Organización",
            "Estado",
            "Cantidad Total",
            "Items",
            "Descripción Items",
            "Usuario",
            "Notas"
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row, transfer in enumerate(transfers, 2):
            worksheet.cell(row=row, column=1, value=transfer.id)
            worksheet.cell(row=row, column=2, value=transfer.fecha_transferencia.strftime("%Y-%m-%d %H:%M:%S") if transfer.fecha_transferencia else "")
            worksheet.cell(row=row, column=3, value=transfer.tipo.value)
            worksheet.cell(row=row, column=4, value=transfer.organizacion_contraparte)
            worksheet.cell(row=row, column=5, value=transfer.estado.value)
            worksheet.cell(row=row, column=6, value=transfer.get_total_quantity())
            worksheet.cell(row=row, column=7, value=transfer.get_total_items())
            
            # Items description
            items_desc = []
            for item in transfer.get_donation_items():
                categoria = item.get('categoria') or item.get('category', 'N/A')
                descripcion = item.get('descripcion') or item.get('description', 'N/A')
                cantidad = item.get('cantidad') or item.get('quantity', 'N/A')
                items_desc.append(f"{categoria}: {descripcion} ({cantidad})")
            
            worksheet.cell(row=row, column=8, value="; ".join(items_desc))
            worksheet.cell(row=row, column=9, value=transfer.usuario_registro or "Sistema")
            worksheet.cell(row=row, column=10, value=transfer.notas or "")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary row at the bottom
        if transfers:
            summary_row = len(transfers) + 3
            worksheet.cell(row=summary_row, column=5, value="TOTAL:")
            worksheet.cell(row=summary_row, column=5).font = Font(bold=True)
            
            total_quantity = sum(t.get_total_quantity() for t in transfers)
            worksheet.cell(row=summary_row, column=6, value=total_quantity)
            worksheet.cell(row=summary_row, column=6).font = Font(bold=True)
            
            total_items = sum(t.get_total_items() for t in transfers)
            worksheet.cell(row=summary_row, column=7, value=total_items)
            worksheet.cell(row=summary_row, column=7).font = Font(bold=True)
    
    def cleanup_expired_files(self):
        """Remove expired Excel files from database and filesystem"""
        with get_db_session() as session:
            expired_files = session.query(ExcelFile).filter(
                ExcelFile.fecha_expiracion < datetime.utcnow()
            ).all()
            
            for excel_file in expired_files:
                # Remove file from filesystem
                try:
                    if os.path.exists(excel_file.ruta_archivo):
                        os.remove(excel_file.ruta_archivo)
                except Exception as e:
                    print(f"Error removing file {excel_file.ruta_archivo}: {e}")
                
                # Remove record from database
                session.delete(excel_file)
            
            session.commit()
            return len(expired_files)