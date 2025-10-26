"""
Excel export service for generating donation reports in Excel format.
"""
import os
import uuid
from typing import List, Optional, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..models.donation import Donation, DonationCategory
from ..models.filter import ExcelFile
from ..models.user import User
from ..services.donation_service import DonationService
from ..utils.database_utils import get_db_session
from ..config import settings


class ExcelExportService:
    """Service for generating Excel exports of donation reports"""
    
    def __init__(self):
        self.donation_service = DonationService()
        self.storage_path = settings.excel_storage_path
        
        # Ensure storage directory exists
        os.makedirs(self.storage_path, exist_ok=True)
    
    def generate_donation_excel(
        self,
        user_id: int,
        categoria: Optional[DonationCategory] = None,
        fecha_desde: Optional[datetime] = None,
        fecha_hasta: Optional[datetime] = None,
        eliminado: Optional[bool] = None,
        user_organization: Optional[str] = None
    ) -> ExcelFile:
        """
        Generate Excel file with donation data filtered by the provided criteria.
        
        Args:
            user_id: ID of the user requesting the export
            categoria: Filter by donation category (optional)
            fecha_desde: Filter donations from this date (optional)
            fecha_hasta: Filter donations until this date (optional)
            eliminado: Filter by eliminated status (optional)
        
        Returns:
            ExcelFile object with file information
        """
        print("[EXCEL] Getting donations from service...")
        # Get filtered donations using the service
        donations = self.donation_service.get_donations_by_filters(
            categoria=categoria,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            eliminado=eliminado,
            user_organization=user_organization
        )
        
        # Pre-load user information to avoid lazy loading issues
        for donation in donations:
            try:
                # Access the relationships to load them
                if donation.usuario_creador:
                    _ = donation.usuario_creador.nombre
                if donation.usuario_modificador:
                    _ = donation.usuario_modificador.nombre
            except:
                pass
        
        print(f"[EXCEL] Got {len(donations)} donations")
        
        # Group donations by category
        print("[EXCEL] Grouping donations by category...")
        donations_by_category = self._group_donations_by_category(donations)
        print(f"[EXCEL] Grouped into {len(donations_by_category)} categories")
        
        # Generate Excel file
        file_id = str(uuid.uuid4())
        filename = self._generate_filename(categoria, fecha_desde, fecha_hasta)
        file_path = os.path.join(self.storage_path, f"{file_id}_{filename}")
        
        workbook = self._create_workbook(donations_by_category)
        workbook.save(file_path)
        
        # Save file record to database
        with get_db_session() as session:
            excel_file = ExcelFile.create_with_expiration(
                usuario_id=user_id,
                nombre_archivo=filename,
                ruta_archivo=file_path,
                hours_to_expire=24  # Files expire after 24 hours
            )
            session.add(excel_file)
            session.commit()
            session.refresh(excel_file)
            
            return excel_file
    
    def _group_donations_by_category(self, donations: List[Donation]) -> Dict[DonationCategory, List[Donation]]:
        """Group donations by category for separate Excel sheets"""
        grouped = {}
        for i, donation in enumerate(donations):
            print(f"[EXCEL] Processing donation {i+1}: ID={donation.id}")
            try:
                categoria = donation.categoria
                print(f"[EXCEL] Donation {i+1} category: {categoria}")
                if categoria not in grouped:
                    grouped[categoria] = []
                grouped[categoria].append(donation)
            except Exception as e:
                print(f"[EXCEL] Error processing donation {i+1}: {e}")
                raise
        return grouped
    
    def _generate_filename(
        self,
        categoria: Optional[DonationCategory] = None,
        fecha_desde: Optional[datetime] = None,
        fecha_hasta: Optional[datetime] = None
    ) -> str:
        """Generate descriptive filename for the Excel export"""
        parts = ["reporte_donaciones"]
        
        if categoria:
            parts.append(f"categoria_{categoria.value.lower()}")
        
        if fecha_desde:
            parts.append(f"desde_{fecha_desde.strftime('%Y%m%d')}")
        
        if fecha_hasta:
            parts.append(f"hasta_{fecha_hasta.strftime('%Y%m%d')}")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        parts.append(timestamp)
        
        return f"{'_'.join(parts)}.xlsx"
    
    def _create_workbook(self, donations_by_category: Dict[DonationCategory, List[Donation]]) -> Workbook:
        """Create Excel workbook with separate sheets for each category"""
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create sheets for each category
        for category, donations in donations_by_category.items():
            sheet_name = self._get_sheet_name(category)
            worksheet = workbook.create_sheet(title=sheet_name)
            self._populate_worksheet(worksheet, donations, category)
        
        # If no donations, create an empty summary sheet
        if not donations_by_category:
            worksheet = workbook.create_sheet(title="Sin Datos")
            worksheet['A1'] = "No se encontraron donaciones con los filtros aplicados"
            worksheet['A1'].font = Font(bold=True)
        
        return workbook
    
    def _get_sheet_name(self, category: DonationCategory) -> str:
        """Get human-readable sheet name for category"""
        category_names = {
            DonationCategory.ROPA: "ROPA",
            DonationCategory.ALIMENTOS: "ALIMENTOS", 
            DonationCategory.JUGUETES: "JUGUETES",
            DonationCategory.UTILES_ESCOLARES: "UTILES_ESCOLARES",
            DonationCategory.MEDICAMENTOS: "MEDICAMENTOS",
            DonationCategory.LIBROS: "LIBROS",
            DonationCategory.ELECTRODOMESTICOS: "ELECTRODOMESTICOS",
            DonationCategory.MUEBLES: "MUEBLES",
            DonationCategory.OTROS: "OTROS"
        }
        return category_names.get(category, category.value)
    
    def _populate_worksheet(self, worksheet, donations: List[Donation], category: DonationCategory):
        """Populate worksheet with donation data - NEW FORMAT per requirements"""
        # Define headers according to requirements: 
        # Fecha de Alta, Descripcion, Cantidad, Eliminado, Usuario Alta, Usuario Modificación
        headers = [
            "Fecha de Alta",
            "Descripción", 
            "Cantidad",
            "Eliminado",
            "Usuario Alta",
            "Usuario Modificación"
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows - only the required columns, no summaries
        for row, donation in enumerate(donations, 2):
            # Fecha de Alta
            worksheet.cell(row=row, column=1, value=donation.fecha_alta.strftime("%Y-%m-%d") if donation.fecha_alta else "")
            
            # Descripción
            worksheet.cell(row=row, column=2, value=donation.descripcion or "")
            
            # Cantidad
            worksheet.cell(row=row, column=3, value=donation.cantidad)
            
            # Eliminado
            worksheet.cell(row=row, column=4, value="Sí" if donation.eliminado else "No")
            
            # Usuario Alta - try to get name, fallback to ID
            usuario_alta = "N/A"
            if donation.usuario_alta:
                try:
                    # Try to get user name from relationship if available
                    if hasattr(donation, 'usuario_creador') and donation.usuario_creador:
                        usuario_alta = f"{donation.usuario_creador.nombre} {donation.usuario_creador.apellido}"
                    else:
                        usuario_alta = f"Usuario ID: {donation.usuario_alta}"
                except:
                    usuario_alta = f"Usuario ID: {donation.usuario_alta}"
            worksheet.cell(row=row, column=5, value=usuario_alta)
            
            # Usuario Modificación - try to get name, fallback to ID
            usuario_modificacion = "N/A"
            if donation.usuario_modificacion:
                try:
                    # Try to get user name from relationship if available
                    if hasattr(donation, 'usuario_modificador') and donation.usuario_modificador:
                        usuario_modificacion = f"{donation.usuario_modificador.nombre} {donation.usuario_modificador.apellido}"
                    else:
                        usuario_modificacion = f"Usuario ID: {donation.usuario_modificacion}"
                except:
                    usuario_modificacion = f"Usuario ID: {donation.usuario_modificacion}"
            worksheet.cell(row=row, column=6, value=usuario_modificacion)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # NO SUMMARY ROW - per requirements, no summaries in Excel
    
    def get_file_by_id(self, file_id: str) -> Optional[ExcelFile]:
        """Get Excel file record by ID"""
        with get_db_session() as session:
            return session.query(ExcelFile).filter(ExcelFile.id == file_id).first()
    
    def cleanup_expired_files(self):
        """Remove expired Excel files from database and filesystem"""
        with get_db_session() as session:
            expired_files = session.query(ExcelFile).filter(
                ExcelFile.fecha_expiracion < datetime.utcnow()
            ).all()
            
            for excel_file in expired_files:
                # Remove file from filesystem
                try:
                    if os.path.exists(excel_file.ruta_archivo):
                        os.remove(excel_file.ruta_archivo)
                except Exception as e:
                    print(f"Error removing file {excel_file.ruta_archivo}: {e}")
                
                # Remove record from database
                session.delete(excel_file)
            
            session.commit()
            return len(expired_files)